# -*- coding: utf-8 -*-
"""
import-ratings.py — מייבא את קובץ העבודה של הליגה (xlsx) אל squads.source.json.

מקור האמת לשווי השחקנים הוא גיליון 'דירוגים'. מקור האמת לסגלים הוא
גיליון 'סגלים'. הסקריפט לא ממציא ערכים: שחקן שלא מדורג מקבל את רצפת
הקובץ (דרג 5 · 1M) ומסומן rated=False כדי שהדוח יראה את זה.

מזהי שחקנים וקבוצות נשמרים יציבים מול הקובץ הקיים — הרכבים שהוגשו
מצביעים על המזהים האלה.
"""
import json, re, unicodedata, collections, sys
from pathlib import Path
import openpyxl

HERE = Path(__file__).parent
XLSX = sys.argv[1]                                   # סגלים (+ דירוגים כגיבוי)
RATINGS_CSV = sys.argv[2] if len(sys.argv) > 2 else None  # דירוגים מעודכנים, גובר
SRC  = HERE / "squads.source.json"

# ---------- נרמול ----------
GERESH = dict.fromkeys(map(ord, "׳״'\"`’‘“”"), None)

def norm(s):
    if s is None: return ""
    s = unicodedata.normalize("NFKC", str(s)).strip()
    s = s.translate(GERESH)
    s = re.sub(r"\(.*?\)", "", s)          # "(עולה חדשה)"
    s = re.sub(r"[֑-ׇ]", "", s)  # ניקוד
    s = re.sub(r"\s+", " ", s).strip()
    return s

def norm_team(s):
    s = norm(s)
    # קריית/קרית, וריאציות איות נפוצות
    s = s.replace("קריית", "קרית").replace("פתח תקוה", "פתח תקווה")
    return s

# תווים ערביים שנשתלו בטעות בתוך שם עברי (החלפת פריסת מקלדת).
# מתקנים במפורש ומדווחים — לא "מנקים בשקט".
ARABIC_SLIP = {
    "ت": "ת", "ا": "א", "ب": "ב", "و": "ו", "ي": "י",
    "أ": "א", "إ": "א", "ة": "ה", "ه": "ה", "م": "מ",
    "ن": "נ", "ل": "ל", "ر": "ר", "س": "ס", "ك": "כ",
    "ح": "ח", "ع": "ע", "ق": "ק", "ف": "פ", "ص": "צ",
    "ط": "ט", "د": "ד", "ز": "ז", "ج": "ג", "ش": "ש",
}
ARABIC_RE = re.compile(r"[\u0600-\u06FF]")
fixed_names = []

def fix_script(s):
    if not s or not ARABIC_RE.search(s):
        return s
    out = "".join(ARABIC_SLIP.get(ch, ch) for ch in s)
    out = re.sub(r"\s+", " ", out).strip()
    fixed_names.append((s, out))
    return out

def norm_name(s):
    s = norm(s)
    s = s.replace("־", " ").replace("-", " ")
    return re.sub(r"\s+", " ", s).strip()

# עמדות: הגיליון משתמש בשפת כדורגל, לא בקודים
POS = {
    "שוער": "GK",
    "בלם": "DEF", "מגן": "DEF", "הגנה": "DEF", "מגן ימני": "DEF", "מגן שמאלי": "DEF",
    "קשר": "MID", "כנף": "MID", "קשר הגנתי": "MID", "קשר התקפי": "MID",
    "חלוץ": "FWD", "חלון/חלוץ": "FWD", "חלוץ מרכזי": "FWD",
}
SRC_POS = {"GK": "GK", "DEF": "DF", "MID": "MF", "FWD": "FW"}

# ---------- קריאה ----------
wb = openpyxl.load_workbook(XLSX, data_only=True)

ratings = {}          # (team, name) -> dict
ratings_by_name = {}  # name -> dict   (גיבוי כשהקבוצה לא תואמת)

def add_rating(overall, tier, price, name, team):
    rec = {"overall": int(overall), "tier": int(tier), "price": int(price),
           "name": norm_name(name), "team": norm_team(team)}
    ratings[(rec["team"], rec["name"])] = rec
    ratings_by_name.setdefault(rec["name"], []).append(rec)

if RATINGS_CSV:
    # קובץ הדירוגים המעודכן גובר. השורות בו לא ממוינות ויש בו כותרת
    # חוזרת באמצע — מסננים לפי "האם העמודה הראשונה היא מספר".
    import csv as _csv
    with open(RATINGS_CSV, encoding="utf-8-sig", newline="") as fh:
        for row in _csv.reader(fh):
            if len(row) < 6 or not str(row[0]).strip().isdigit():
                continue
            add_rating(row[0], row[2], row[3], row[4], row[5])
else:
    for r in wb["דירוגים"].iter_rows(min_row=4, values_only=True):
        if r[0] is None: continue
        add_rating(r[0], r[2], r[3], r[4], r[5])

# עמדות ידועות מהתוצר האחרון שנוצר (src/data/squads.ts) — רשת ביטחון
# לשורות משובשות באקסל. נקרא משם ולא מ-squads.source.json כדי שהוא
# ישרוד גם הרצה חוזרת של הסקריפט.
PREV_POS = {}
_gen = HERE.parent / "src" / "data" / "squads.ts"
if _gen.exists():
    _txt = _gen.read_text("utf-8")
    _teamname = {m.group(1): norm_team(m.group(2)) for m in re.finditer(
        r'"id":\s*"(T\d+)",\s*"externalId":[^,]+,\s*"nameHe":\s*"([^"]+)"', _txt)}
    for m in re.finditer(
        r'"teamId":\s*"(T\d+)",\s*"position":\s*"(\w+)",\s*"nameHe":\s*"([^"]+)"', _txt):
        tid, pos, nm = m.groups()
        if tid in _teamname:
            PREV_POS[(_teamname[tid], norm_name(nm))] = {"GK":"GK","DEF":"DF","MID":"MF","FWD":"FW"}[pos]

squads = collections.OrderedDict()
skipped, recovered = [], []
for r in wb["סגלים"].iter_rows(min_row=4, values_only=True):
    if not r[0]: continue
    team, name, pos, number, nat, cap = r[:6]
    p = POS.get(norm(pos))
    if p is None:
        # שורה משובשת באקסל (הזזת עמודות). לא מנחשים ולא מוחקים שחקן:
        # נופלים בחזרה לעמדה הידועה מהגרסה הקודמת, ומדווחים.
        prev = PREV_POS.get((norm_team(team), norm_name(name)))
        if prev:
            p = {"GK": "GK", "DF": "DEF", "MF": "MID", "FW": "FWD"}[prev]
            recovered.append((norm_team(team), norm_name(name), prev))
        else:
            skipped.append((norm_team(team), norm_name(name), norm(pos)))
            continue
    squads.setdefault(norm_team(team), []).append({
        "name": norm_name(name), "name_raw": fix_script(str(name).strip()), "pos": p,
        "number": int(number) if number and str(number).isdigit() else None,
        "nat": norm(nat) or None, "captain": bool(cap),
    })

# ---------- שמירת מזהים יציבים ----------
old = json.loads(SRC.read_text("utf-8"))
old_team_id, old_player_id = {}, {}
for t in old["teams"]:
    old_team_id[norm_team(t["name_he"])] = t["team_id"]
    for p in t["players"]:
        _k = (norm_team(t["name_he"]), norm_name(p["name_he"]), p["position"])
        old_player_id[_k] = p["id"]
old_meta = {norm_team(t["name_he"]): t for t in old["teams"]}

next_team = max(old_team_id.values(), default=0) + 1
used_pid = set(old_player_id.values())
def new_pid(tid):
    """מזהה יציב וייחודי. הבלוק tid*100 נשמר לקריאוּת, אבל אם הוא מלא
    גולשים לטווח 90000+ במקום לדרוס מזהה קיים של קבוצה אחרת."""
    base = tid * 100 + 1
    limit = tid * 100 + 100
    while base in used_pid and base < limit:
        base += 1
    if base >= limit:
        base = 90001
        while base in used_pid:
            base += 1
    used_pid.add(base)
    return base

# ---------- הרכבה ----------
used_ratings, fallback_used = set(), []
teams_out, report = [], {"matched": 0, "unrated": 0, "name_only": 0, "skipped": skipped, "recovered": recovered}
for team_he, players in squads.items():
    tid = old_team_id.get(team_he)
    if tid is None:
        tid = next_team; next_team += 1
    meta = old_meta.get(team_he, {})
    out = []
    for p in players:
        rec = ratings.get((team_he, p["name"]))
        if rec is None:
            cands = ratings_by_name.get(p["name"])
            if cands and len(cands) == 1:
                rec = cands[0]
                report["name_only"] += 1
                fallback_used.append((team_he, p["name"], cands[0]["team"]))
        if rec is not None:
            used_ratings.add((rec["team"], rec["name"]))
        if rec:
            tier, price, overall = rec["tier"], rec["price"], rec["overall"]
            report["matched"] += 1
        else:
            tier, price, overall = 5, 1, None   # רצפת הקובץ
            report["unrated"] += 1
        # שני שחקנים באותה קבוצה יכולים לשאת אותו שם (ליאם כהן MF/FW).
        # המפתח כולל עמדה, וכל מזהה נתפס פעם אחת בלבד.
        _key = (team_he, p["name"], SRC_POS[p["pos"]])
        pid = old_player_id.get(_key)
        if pid is None or pid in used_pid:
            pid = new_pid(tid)
        else:
            used_pid.add(pid)
        out.append({
            "id": pid, "name_he": p["name_raw"], "name_en": "TBD",
            "position": SRC_POS[p["pos"]], "number": p["number"],
            "tier": tier, "price": price,
            "overall_rank": overall, "rated": rec is not None,
            "nationality": p["nat"], "is_club_captain": p["captain"],
        })
    teams_out.append({
        "team_id": tid, "name_he": team_he,
        "name_en": meta.get("name_en", "TBD"), "short": meta.get("short", team_he[:3]),
        "city": meta.get("city"), "stadium": meta.get("stadium"),
        "players": out,
    })

teams_out.sort(key=lambda t: t["team_id"])
out = {"league": old["league"], "teams": teams_out}
SRC.write_text(json.dumps(out, ensure_ascii=False, indent=2), "utf-8")

print(f"teams={len(teams_out)} players={sum(len(t['players']) for t in teams_out)}")
print(f"matched={report['matched']} (name-only fallback={report['name_only']}) unrated={report['unrated']}")
print("recovered rows:", report["recovered"])
print("script-fixed names:", fixed_names)
print("skipped rows:", report["skipped"])
print("name-only fallbacks:", fallback_used)
_orphans = [f'{k[1]} ({k[0]})' for k in ratings if k not in used_ratings]
print("ratings with no squad player:", _orphans)
# נכתב לקובץ כדי שיופיע בדוח איכות הנתונים ולא ייעלם בטרמינל
(HERE / "ratings.orphans.json").write_text(json.dumps(_orphans, ensure_ascii=False, indent=2), "utf-8")
dist = collections.Counter(p["price"] for t in teams_out for p in t["players"])
print("price dist:", dict(sorted(dist.items())))
pos = collections.Counter(p["position"] for t in teams_out for p in t["players"])
print("positions:", dict(pos))
